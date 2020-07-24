import openpyxl
from openpyxl.styles import Font
wb = openpyxl.Workbook()
sheet = wb['Sheet']
italic24Font = Font(italic=True, size=24)
sheet['A1'].font = italic24Font
sheet['A1'] = 'Hello world!'
sheet['A2'] = 'Hello world!'
wb.save('styled.xlsx')

# Running Formulas
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)'
wb.save('writeFormula.xlsx')

# Reading Formulas
wbFormulas = openpyxl.load_workbook('writeFormula.xlsx')
sheet = wbFormulas.active
print(sheet['A3'].value)

wbDataOnly = openpyxl.load_workbook('writeFormula.xlsx', data_only=True)
sheet = wbDataOnly.active
print(sheet['A3'].value)

# Setting Row Height and Column Width
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
wb.save('dimensions.xlsx')

# Merging and Unmerging Cells
wb = openpyxl.Workbook()
sheet = wb.active
sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged together.'
sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two merged cells.'
# sheet.unmerge_cells('A1:D3')
# sheet.unmerge_cells('C5:D5')
wb.save('merged.xlsx')

# Freeze Panes
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active
sheet.freeze_panes = 'A2'
wb.save('freezeExample.xlsx')

# Charts *This is deprecated
wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1,11):	# create some data in column A:
	sheet['A' + str(i)] = i
refObj = openpyxl.charts.Reference(sheet, (1, 1), (10, 1))
seriesObj = openpyxl.charts.Series(refObj, title='First series')
chartObj = openpyxl.charts.BarChart()
chartObj.append(seriesObj)
chartObj.drawing.top = 50 # set the position
chartObj.drawing.left = 100
chartObj.drawing.width = 300 # set the size
chartObj.drawing.height = 200

sheet.add_chart(chartObj)
wb.save('sampleChart.xlsx')