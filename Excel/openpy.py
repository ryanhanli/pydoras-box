import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('example.xlsx')
print(type(wb))
print(wb.sheetnames)
print(wb['Sheet3'])
print(wb.active)
sheet = wb['Sheet1']
print(sheet['A1'].value)
c = sheet['B1']
print(c.value)

print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)
for i in range(1,8,2):
	print(i, sheet.cell(row=i, column=2).value)
print(sheet.max_row)
print(sheet.max_column)

print(get_column_letter(900))
print(column_index_from_string('AA'))

for rowOfCellObjects in sheet['A1':'C3']:
	for cellObj in rowOfCellObjects:
		print(cellObj.coordinate, cellObj.value)
	print('--- END OF ROW ---')
	
print(tuple(sheet.columns))
for cellObj in (tuple(sheet.columns)[1]):
	print(cellObj.value)