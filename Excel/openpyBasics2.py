import openpyxl

#Create a new workbook
wb = openpyxl.Workbook()
print(wb.sheetnames)
#['Sheet']
sheet = wb.active
sheet.title='Spam Bacon Eggs Sheet'
#It doesn't change until you save it
# wb = openpyxl.load_workbook('example.xlsx')

#wb.save('whatevernameyouwant.xlsx')

# Creating and Removing Sheets
wb.create_sheet(title='Sheet1') # Returns Worksheet object Sheet1
print(wb.sheetnames)
wb.create_sheet(index=0, title='First Sheet')
print(wb.sheetnames)
wb.create_sheet(index=2, title='Middle Sheet')
print(wb.sheetnames)
wb.remove(wb['Middle Sheet'])
wb.remove(wb['Sheet1'])
print(wb.sheetnames)

# Writing values to cells